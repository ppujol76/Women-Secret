import PySimpleGUI as sg
import pandas as pd
import numpy as np
import os
import openpyxl 
from datetime import date

def main():
    # Define the layout of the GUI
    layout = [
        [sg.Text("Select a file:")],
        [sg.InputText(key="file_path"), sg.FileBrowse()],
        [sg.Button("Upload", key="upload_button")],
        [sg.Button("Procesar", key="process_button")]
    ]

    # Create the window
    window = sg.Window("File Upload Example", layout)

    while True:
        event, values = window.read()

        if event == sg.WINDOW_CLOSED:
            break
        elif event == "upload_button":
            file_path = values["file_path"]
            if file_path:
                sg.popup(f"Selected file: {file_path}")
                print("Reading file:", file_path)
            else:
                sg.popup("No file selected!")
        elif event == "process_button":
            print("Processament del fitxer iniciat")
            xls = pd.ExcelFile(file_path)
            print(os.getcwd())
            print(xls.sheet_names)

            df1 = xls.parse('stock', dtype={"PATRON CD":str}, nrows=93, usecols=['ARTICLE',  'DESCRIPCIO', 'PRESENTACIO', 'TALLA', 'COLOR', 'UNITATS VENUDES', 'PATRON CD', 'Unnamed: 9', 'Unnamed: 10'])
            df1.rename(columns={'Unnamed: 9': 'Q_caja_1', 'Unnamed: 10': 'Q_caja_2'}, inplace=True)
            df1_ok = df1.dropna(subset=["UNITATS VENUDES"])
            df1_ok = df1_ok[df1_ok["UNITATS VENUDES"]>0]
            print("Total d'unitats venudes:", df1_ok["UNITATS VENUDES"].sum())

            first_time = True
            while True:
                if first_time == True:
                    caixes_petites = input("Indica el tipus de caixa que utilitzarem, de 72 o de 80?\n")
                    first_time = False
                else:
                    caixes_petites = input("Ha de indicar el número 72 o el 80\n")
                if caixes_petites in ["72", "80"]:
                    caixes_petites = int(caixes_petites)
                    break

            mides_caixes = {caixes_petites:(72, 80), 150:(150, 150), 198:(198, 198)}
            dfs = {}

            for mida in mides_caixes.keys():
                tamany_caixa = mida
                df_aux = df1_ok[(df1_ok["Q_caja_1"] == mides_caixes[mida][0]) | (df1_ok["Q_caja_1"] == mides_caixes[mida][1])]
                df_aux2 = pd.DataFrame(columns=["Nº CAJA", "ARTICULO", "COLOR", "TALLA", "UNIDADES", "", "TOTAL UDS", "TIPO CAJA"])

                indice = 1
                for i in df_aux.index:
#                    print(tamany_caixa)
                    num_caixes = int(df_aux.loc[i,"UNITATS VENUDES"] // tamany_caixa)
                    pico = int(df_aux.loc[i,"UNITATS VENUDES"] % tamany_caixa)
                    article, talla, color = df_aux.loc[i,"ARTICLE"], df_aux.loc[i,"TALLA"], df_aux.loc[i,"COLOR"]
                #  print(f"Article:{article}, TALLA:{talla}, COLOR:{color}, Num Caixes plenes:{num_caixes}, Pico:{pico}")

                    for caixa in range(num_caixes):
                #    print(indice)
                        df_aux2 = pd.concat([df_aux2, pd.DataFrame({"Nº CAJA":[indice], "ARTICULO":[article], "COLOR":[color], "TALLA":[talla], "UNIDADES":[tamany_caixa], "TIPO CAJA":0})])
                        #df_aux2.append({"Nº CAJA":indice, "ARTICULO":article},ignore_index=True)    ##, "COLOR":color, "TALLA":talla, "UNIDADES":tamany_caixa}, ignore_index=True)
                        indice += 1
                    if pico != 0:

                        if (tamany_caixa==72 or tamany_caixa==80):

                    ##=SI(F8>60;"0";SI(F8>52;"5";SI(F8>48;"4";SI(F8>40;"3";SI(F8>30;"2";SI(F8>25;"1";SI(F8>15;"25";SI(F8>1;"30";""))))))))
                            condicions = [
                                pico>60,
                                np.logical_and(pico>52,pico<=60),
                                np.logical_and(pico>48, pico<=52),
                                np.logical_and(pico>40, pico<=48),
                                np.logical_and(pico>30,pico<=40),
                                np.logical_and(pico>25, pico<=30),
                                np.logical_and(pico>15, pico<=25),
                                np.logical_and(pico>1, pico<=15),
                                ]
                            valors = [0, 5, 4, 3, 2, 1, 25, 30]
                            tipo_caja = np.select(condicions, valors)

                        elif (tamany_caixa==150):
                    ## =SI(F17>103;"0";SI(F17>97;"5";SI(F17>87;"4";SI(F17>77;"3";SI(F17>59;"2";SI(F17>50;"1";SI(F17>35;"25";SI(F17>1;"30";""))))))))
                            condicions = [
                                pico>103,
                                np.logical_and(pico>97,pico<=103),
                                np.logical_and(pico>87, pico<=97),
                                np.logical_and(pico>77, pico<=87),
                                np.logical_and(pico>59,pico<=77),
                                np.logical_and(pico>50, pico<=59),
                                np.logical_and(pico>35, pico<=50),
                                np.logical_and(pico>1, pico<=35),
                                ]
                            valors = [0, 5, 4, 3, 2, 1, 25, 30]
                            tipo_caja = np.select(condicions, valors)

                        elif tamany_caixa==198:
                            condicions = [
                                pico>165,
                                np.logical_and(pico>159,pico<=165),
                                np.logical_and(pico>125, pico<=159),
                                np.logical_and(pico>108, pico<=125),
                                np.logical_and(pico>79,pico<=108),
                                np.logical_and(pico>50, pico<=79),
                                np.logical_and(pico>32, pico<=50),
                                np.logical_and(pico>1, pico<=32),
                                ]
                            valors = [0, 5, 4, 3, 2, 1, 25, 30]
                            tipo_caja = np.select(condicions, valors)


                        df_aux2 = pd.concat([df_aux2, pd.DataFrame({"Nº CAJA":[indice], "ARTICULO":[article], "COLOR":[color], "TALLA":[talla], "UNIDADES":[pico], "TOTAL UDS":[num_caixes * tamany_caixa + pico], "TIPO CAJA":tipo_caja})])
                        indice += 1
                    else:
                        df_aux2.iloc[len(df_aux2)-1, 6] = num_caixes * tamany_caixa + pico

                dfs[mida]= df_aux2.copy()
                dfs[mida].to_excel(f"./packings_{mida}.xlsx", index=False)

                wb = openpyxl.load_workbook(f"./packings_{mida}.xlsx")
                ws = wb.active    #   o també ws = wb['Sheet1']

                ws.insert_rows(1)

                current_date = date.today()
                current_date = current_date.strftime('%d-%m-%Y')

                ws['A1'] = "SALIDA:"
                ws['B1'] = "WOMEN SECRET"
                ws['D1'] = ws.max_row - 2
                ws['E1'] = f"=SUM(E3:E{ws.max_row})"
                ws['F1'] = "FECHA:"
                ws['G1'] = current_date 
                wb.save(f"./packings_{mida}_definitiu.xlsx")

            print("Els Excels s'han generat correctament")
    window.close()

if __name__ == "__main__":
    main()